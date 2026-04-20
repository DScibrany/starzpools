#!/usr/bin/env python3
"""Daily update pipeline for the STARZ Pasienky dashboard.

The pool pages on bratislava.sk expose two artefacts under the
"Otváracie hodiny a vyťaženosť bazéna" section:

  * "Časový rozpis voľných plaveckých dráh pre verejnosť" — SharePoint
    link that serves an XLSX workbook with 14 days of 15-min lane data.
  * Cenník — a PDF on bratislavask.s3.bratislava.sk.

Both URLs rotate (SharePoint tokens, S3 object IDs), so the script
discovers the current ones by scraping the public page.

Commands:
  fetch       — download XLSX + PDF from the public pages into data/starz/
  transform   — read local XLSX files, write schedule*.json
  pricing     — compare downloaded cenník to the stored copy, update
                pricing.json.status accordingly
  update      — fetch, transform, pricing (what the daily cron runs)
"""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import io
import json
import os
import re
import sys
import time
import urllib.parse
from pathlib import Path
from typing import Iterable

try:
    import requests
except ImportError:
    requests = None  # transform-only runs do not need it

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

import openpyxl

try:
    import xlrd  # legacy .xls support; SharePoint serves that format now
except ImportError:
    xlrd = None

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data" / "starz"

POOL_PAGES = {
    "25m": "https://bratislava.sk/vzdelavanie-a-volny-cas/starz/prevadzky-sportoviska/mestska-plavaren-pasienky-25m",
    "50m": "https://bratislava.sk/vzdelavanie-a-volny-cas/starz/prevadzky-sportoviska/mestska-plavaren-pasienky-50m",
}
XLSX_FILES = {
    "25m": DATA_DIR / "2026-Rezervacie-MPP25-Verejnost.xlsx",
    "50m": DATA_DIR / "2026-Rezervacie-MPP50-Verejnost.xlsx",
}
POOL_META = {
    "25m": {
        "out": ROOT / "schedule.json",
        "name": "Mestská plaváreň Pasienky 25 m",
        "max_lanes": 4,
    },
    "50m": {
        "out": ROOT / "schedule-50m.json",
        "name": "Mestská plaváreň Pasienky 50 m",
        "max_lanes": 8,
    },
}
PRICING_PDF = DATA_DIR / "pricing-current.pdf"
PRICING_JSON = ROOT / "pricing.json"

SCHEDULE_LINK_TEXT = "Časový rozpis voľných plaveckých dráh"
PRICING_LINK_TEXT = "Cenník"

SLOT_MINUTES = 15
DAY_START_HOUR = 5
DAY_END_HOUR = 24
SLOTS_PER_DAY = (DAY_END_HOUR - DAY_START_HOUR) * 60 // SLOT_MINUTES  # 76
DATA_COL_START = 4  # column E (0-based) in the Verejnost sheet
ROWS_PER_DAY = 9    # date row + 7 service rows + "Počet voľných dráh" row
COUNTER_ROW_OFFSET = 8
SK_WEEKDAYS = {
    0: "pondelok", 1: "utorok", 2: "streda", 3: "štvrtok",
    4: "piatok", 5: "sobota", 6: "nedeľa",
}


# ---------------------------------------------------------------- scraping --

def _require(module, name: str):
    if module is None:
        raise SystemExit(f"Missing dependency '{name}'. Run: pip install {name}")


# Browser-like defaults. bratislava.sk's WAF rejects UAs that self-identify
# as bots ("starzpools-bot" → 403), so we pose as a recent desktop Chrome and
# send the Accept/Accept-Language a real browser would. Kept centralised so
# both the HTML scrape and the XLSX/PDF downloads use the same identity.
_DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,image/apng,*/*;q=0.8"
    ),
    "Accept-Language": "sk-SK,sk;q=0.9,en;q=0.8",
    "Referer": "https://bratislava.sk/",
}

_RETRY_STATUSES = {408, 425, 429, 500, 502, 503, 504}
_HTTP_ATTEMPTS = 4


def _http_get(url: str, **kw) -> "requests.Response":
    _require(requests, "requests")
    headers = dict(_DEFAULT_HEADERS)
    headers.update(kw.pop("headers", {}))

    last_err: Exception | None = None
    for attempt in range(1, _HTTP_ATTEMPTS + 1):
        try:
            resp = requests.get(
                url, headers=headers, timeout=60, allow_redirects=True, **kw
            )
        except requests.exceptions.RequestException as e:
            last_err = e
        else:
            if resp.status_code < 400:
                return resp
            last_err = requests.HTTPError(
                f"{resp.status_code} {resp.reason} for {url}", response=resp
            )
            # 4xx other than the explicit retry set are not worth retrying.
            if resp.status_code not in _RETRY_STATUSES:
                break
        if attempt < _HTTP_ATTEMPTS:
            time.sleep(2 ** (attempt - 1))  # 1, 2, 4 s
    raise RuntimeError(f"GET {url} failed after {_HTTP_ATTEMPTS} attempts: {last_err}")


def _href_interesting(href: str) -> bool:
    """Heuristic for anchors worth logging when discovery fails — filters out
    nav / utility links so the error preview shows file links only."""
    h = href.lower()
    return any(
        tok in h
        for tok in (
            "sharepoint", "onedrive", "dropbox",
            "s3.bratislava.sk", "s3.amazonaws", "cdn.",
            ".xlsx", ".xls", ".pdf", ".docx",
            "rezerv", "rozpis", "verejnost", "cennik",
        )
    )


def _matches_xlsx(text_low: str, href_low: str) -> bool:
    """STARZ schedule workbook candidate. SharePoint URLs are opaque
    (`/:x:/t/<site>/IQ...`) and the anchor text is now a generic
    "Prejsť na stránku", so the strongest signal is the href: either a
    direct `.xlsx` or a SharePoint Excel share (`/:x:/`)."""
    if "rozpis" in text_low and (
        "voľn" in text_low or "plavec" in text_low or "dráh" in text_low
    ):
        return True
    if ".xlsx" in href_low:
        return True
    if "sharepoint.com" in href_low and "/:x:/" in href_low:
        return True
    return False


def _matches_pdf(text_low: str, href_low: str) -> bool:
    """Cenník PDF. Tolerate both diacritic spellings and the lowercase 'cennik'
    that appears in S3 object keys."""
    looks_like_pdf = (
        href_low.endswith(".pdf")
        or ".pdf?" in href_low
        or "s3.bratislava.sk" in href_low
    )
    if not looks_like_pdf:
        return False
    return (
        text_low.startswith("cenník")
        or "cenník" in text_low
        or "cennik" in text_low
        or "cennik" in href_low
    )


def discover_links(page_url: str) -> dict:
    """Return {'xlsx_urls': [url, ...], 'pdf': url | None,
    'pdf_text': str | None, 'candidates': [...]}.

    The page lists multiple SharePoint workbooks (25 m + 50 m on the same
    page) behind generic "Prejsť na stránku" buttons, so the anchor text is
    no longer a reliable discriminator. Return every XLSX-looking URL we
    find; the caller picks the right one by reading the workbook's title.
    """
    _require(BeautifulSoup, "beautifulsoup4")
    html = _http_get(page_url).text
    soup = BeautifulSoup(html, "html.parser")
    xlsx_urls: list[str] = []
    pdf_url = None
    pdf_text = None
    candidates = []
    for a in soup.find_all("a", href=True):
        text = " ".join(a.get_text(" ", strip=True).split())
        href = a["href"]
        if not href or href.startswith("#") or href.startswith("javascript:"):
            continue
        abs_url = urllib.parse.urljoin(page_url, href)
        text_low = text.lower()
        href_low = abs_url.lower()
        candidates.append({"text": text, "href": abs_url})
        if _matches_xlsx(text_low, href_low) and abs_url not in xlsx_urls:
            xlsx_urls.append(abs_url)
        if pdf_url is None and _matches_pdf(text_low, href_low):
            pdf_url = abs_url
            pdf_text = text
    return {
        "xlsx_urls": xlsx_urls,
        "pdf": pdf_url,
        "pdf_text": pdf_text,
        "candidates": candidates,
    }


def _download_binary(url: str, dest: Path) -> Path:
    resp = _http_get(url, stream=True)
    dest.parent.mkdir(parents=True, exist_ok=True)
    with dest.open("wb") as f:
        for chunk in resp.iter_content(chunk_size=65536):
            if chunk:
                f.write(chunk)
    return dest


# Row-2 of the Verejnost sheet contains a human-readable title, e.g.
# "Otváracie hodiny pre verejnosť na 25 m bazéne …". The pool is the only
# reliable discriminator between the two workbooks linked on each page.
_POOL_TITLE_MARKER = {"25m": "25 m bazén", "50m": "50 m bazén"}


def _sharepoint_download_candidates(url: str) -> list[str]:
    """Return URL variants that might yield the raw XLSX for a SharePoint
    share link.  Team-site share URLs of the form
    ``https://<tenant>.sharepoint.com/:x:/t/<site>/<token>?e=…`` ignore
    ``?download=1`` and serve the HTML viewer, but the same token works
    against ``/sites/<site>/_layouts/15/download.aspx?share=<token>`` and
    the equivalent ``guestaccess.aspx`` endpoint.
    """
    if "sharepoint.com" not in url.lower():
        return [url]

    candidates: list[str] = []
    if re.search(r"[?&]download=1(?:&|$)", url):
        candidates.append(url)
    else:
        sep = "&" if "?" in url else "?"
        candidates.append(f"{url}{sep}download=1")

    m = re.match(
        r"(https?://[^/]+)/:x:/[tg]/([^/]+)/([^/?#]+)", url
    )
    if m:
        host, site, token = m.groups()
        candidates.append(
            f"{host}/sites/{site}/_layouts/15/download.aspx?share={token}"
        )
        candidates.append(
            f"{host}/sites/{site}/_layouts/15/guestaccess.aspx?"
            f"share={token}&download=1"
        )
    return candidates


# XLSX files are zip containers; legacy XLS files are OLE2 compound docs.
# SharePoint serves the STARZ workbooks as .xls (Content-Type
# application/vnd.ms-excel), so accept either magic.
_XLSX_MAGIC = b"PK\x03\x04"
_XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _looks_like_workbook(content: bytes) -> bool:
    return content.startswith(_XLSX_MAGIC) or content.startswith(_XLS_MAGIC)


def _fetch_xlsx_bytes(url: str) -> bytes:
    """Try each SharePoint download variant until one returns something that
    starts with an Excel workbook magic (.xls OLE2 or .xlsx zip). Raises
    with per-attempt details otherwise so the Actions log points at the
    failing endpoint."""
    errors: list[str] = []
    for candidate in _sharepoint_download_candidates(url):
        try:
            resp = _http_get(candidate)
        except Exception as e:
            errors.append(f"{candidate} -> {type(e).__name__}: {e}")
            continue
        content = resp.content
        if _looks_like_workbook(content):
            return content
        ct = resp.headers.get("Content-Type", "?")
        errors.append(
            f"{candidate} -> HTTP {resp.status_code} {len(content)}B "
            f"Content-Type={ct!r} prefix={content[:16]!r}"
        )
    raise RuntimeError(
        f"No SharePoint variant returned a workbook for {url}; tried:\n  "
        + "\n  ".join(errors)
    )


def _read_verejnost_rows(workbook_bytes: bytes) -> list[list]:
    """Return every row of the 'Verejnost' sheet as a list of lists,
    regardless of whether the workbook is .xlsx (zip) or legacy .xls
    (OLE2). Dates come back as ``datetime`` objects in both branches.
    """
    if workbook_bytes.startswith(_XLSX_MAGIC):
        wb = openpyxl.load_workbook(
            io.BytesIO(workbook_bytes), data_only=True, read_only=True
        )
        if "Verejnost" not in wb.sheetnames:
            raise RuntimeError("xlsx workbook has no 'Verejnost' sheet")
        ws = wb["Verejnost"]
        return [list(row) for row in ws.iter_rows(values_only=True)]

    if workbook_bytes.startswith(_XLS_MAGIC):
        _require(xlrd, "xlrd")
        wb = xlrd.open_workbook(file_contents=workbook_bytes)
        if "Verejnost" not in wb.sheet_names():
            raise RuntimeError("xls workbook has no 'Verejnost' sheet")
        sh = wb.sheet_by_name("Verejnost")
        rows: list[list] = []
        for r in range(sh.nrows):
            row: list = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    tup = xlrd.xldate_as_tuple(val, wb.datemode)
                    val = dt.datetime(*tup)
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    val = None
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    val = bool(val)
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    val = None
                row.append(val)
            rows.append(row)
        return rows

    raise RuntimeError(
        f"Unrecognised workbook magic: {workbook_bytes[:8]!r}"
    )


def _workbook_pool_tag(workbook_bytes: bytes) -> str | None:
    """Return '25m' / '50m' if the workbook's title matches, else None."""
    try:
        rows = _read_verejnost_rows(workbook_bytes)
    except Exception:
        return None
    for row in rows[:6]:
        for value in row:
            if not isinstance(value, str):
                continue
            low = value.lower()
            for tag, marker in _POOL_TITLE_MARKER.items():
                if marker in low:
                    return tag
    return None


def fetch_sources() -> dict:
    """Download XLSX for both pools and the current pricing PDF.

    Returns a dict with URLs found and file paths written.
    """
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    report = {"pools": {}, "pricing": {}}

    # Each pool page lists multiple SharePoint workbooks (25 m + 50 m);
    # download each candidate once and key the cache by URL, then map pool→URL
    # by reading the workbook title.
    xlsx_by_pool: dict[str, str] = {}
    xlsx_cache: dict[str, bytes] = {}
    pool_page_info: dict[str, dict] = {}

    for pool, page_url in POOL_PAGES.items():
        info = discover_links(page_url)
        pool_page_info[pool] = info
        if not info["xlsx_urls"]:
            cands = info.get("candidates", [])
            interesting = [c for c in cands if _href_interesting(c["href"])]
            to_show = interesting if interesting else cands
            formatted = [f"{c['text']!r} -> {c['href']}" for c in to_show[:40]]
            raise RuntimeError(
                f"No XLSX-looking links on {page_url}. "
                f"Saw {len(cands)} anchors, {len(interesting)} with interesting href. "
                f"Showing up to 40:\n  " + "\n  ".join(formatted)
            )
        for url in info["xlsx_urls"]:
            if url in xlsx_cache:
                continue
            try:
                data = _fetch_xlsx_bytes(url)
            except Exception as e:
                # Full per-attempt breakdown is already in the exception body.
                for line in str(e).splitlines():
                    print(f"[fetch] {line}", file=sys.stderr)
                continue
            tag = _workbook_pool_tag(data)
            if tag is None:
                print(
                    f"[fetch] {url}: downloaded {len(data)}B but no "
                    f"'Verejnost' sheet / pool-title marker found",
                    file=sys.stderr,
                )
                continue
            xlsx_cache[url] = data
            # First workbook wins for each pool; duplicates are ignored.
            xlsx_by_pool.setdefault(tag, url)

    missing = [p for p in POOL_PAGES if p not in xlsx_by_pool]
    if missing:
        tried = []
        for pool, info in pool_page_info.items():
            tried.append(f"{pool} page saw: {info['xlsx_urls']}")
        raise RuntimeError(
            f"Could not identify workbook(s) for pool(s): {missing}. "
            + " | ".join(tried)
        )

    for pool, url in xlsx_by_pool.items():
        XLSX_FILES[pool].parent.mkdir(parents=True, exist_ok=True)
        XLSX_FILES[pool].write_bytes(xlsx_cache[url])
        page_info = pool_page_info.get(pool, {})
        report["pools"][pool] = {
            "xlsx": url,
            "pdf": page_info.get("pdf"),
            "pdf_text": page_info.get("pdf_text"),
            "xlsx_candidates": page_info.get("xlsx_urls", []),
        }

    # Pricing PDF — discover on the 25m page (same PDF on both).
    pdf_info = pool_page_info.get("25m", {})
    pdf_url = pdf_info.get("pdf")
    report["pricing"]["url"] = pdf_url
    report["pricing"]["source_page"] = POOL_PAGES["25m"]
    if pdf_url:
        _download_binary(pdf_url, PRICING_PDF)
        report["pricing"]["downloaded"] = True
    else:
        report["pricing"]["downloaded"] = False
    return report


# ----------------------------------------------------------- xlsx transform -

def transform_xlsx(xlsx_path: Path, meta: dict, source_page: str) -> dict:
    rows = _read_verejnost_rows(xlsx_path.read_bytes())
    days = []
    # 0-based indexing; the first date used to land on 1-based row 6.
    date_idx = 5
    while date_idx < len(rows):
        first_col = rows[date_idx][0] if rows[date_idx] else None
        if not isinstance(first_col, dt.datetime):
            break
        counter_idx = date_idx + COUNTER_ROW_OFFSET
        if counter_idx >= len(rows):
            raise RuntimeError(
                f"Workbook truncated: expected 'Počet voľných dráh' row at "
                f"index {counter_idx} but sheet only has {len(rows)} rows"
            )
        counter_row = rows[counter_idx]
        label = counter_row[0] if counter_row else None
        if label != "Počet voľných dráh":
            raise RuntimeError(
                f"Unexpected layout: row {counter_idx + 1} col A = {label!r}, "
                "expected 'Počet voľných dráh'"
            )
        slots = counter_row[DATA_COL_START: DATA_COL_START + SLOTS_PER_DAY]
        # Pad if the legacy .xls export trims trailing empty cells.
        if len(slots) < SLOTS_PER_DAY:
            slots = list(slots) + [None] * (SLOTS_PER_DAY - len(slots))
        free = [int(v) if isinstance(v, (int, float)) else 0 for v in slots]
        days.append({
            "date": first_col.strftime("%Y-%m-%d"),
            "weekday": SK_WEEKDAYS[first_col.weekday()],
            "free": free,
        })
        date_idx += ROWS_PER_DAY

    if not days:
        raise RuntimeError(f"No day rows found in {xlsx_path}")

    return {
        "pool": meta["name"],
        "source": source_page,
        "updated": dt.date.today().strftime("%Y-%m-%d"),
        "note": (
            f"Hodnoty = počet voľných dráh pre verejnosť v danom 15-min bloku "
            f"(max {meta['max_lanes']}). Zdroj: STARZ tabuľka."
        ),
        "timezone": "Europe/Bratislava",
        "slotMinutes": SLOT_MINUTES,
        "dayStart": f"{DAY_START_HOUR:02d}:00",
        "dayEnd": f"{DAY_END_HOUR:02d}:00",
        "maxLanes": meta["max_lanes"],
        "days": days,
    }


def write_schedules() -> list[Path]:
    written = []
    for pool, meta in POOL_META.items():
        xlsx = XLSX_FILES[pool]
        if not xlsx.exists():
            print(f"[transform] skip {pool}: {xlsx} missing", file=sys.stderr)
            continue
        data = transform_xlsx(xlsx, meta, POOL_PAGES[pool])
        meta["out"].write_text(
            json.dumps(data, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        print(f"[transform] wrote {meta['out'].relative_to(ROOT)} "
              f"({len(data['days'])} days)")
        written.append(meta["out"])
    return written


# --------------------------------------------------------------- pricing --

def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def update_pricing_status(pdf_info: dict | None = None) -> dict:
    """Compare the stored PDF to the freshly downloaded one.

    If pdf_info is not supplied, only the stored PDF's fingerprint is
    recorded. Expected keys: url, source_page, downloaded.
    """
    doc = json.loads(PRICING_JSON.read_text(encoding="utf-8"))
    status = doc.get("status", {})
    stored_hash = _sha256(PRICING_PDF) if PRICING_PDF.exists() else None
    status["storedSha256"] = stored_hash
    status["lastChecked"] = dt.datetime.utcnow().strftime("%Y-%m-%dT%H:%MZ")

    if pdf_info is not None:
        status["currentUrl"] = pdf_info.get("url")
        status["sourcePage"] = pdf_info.get("source_page")
        if not pdf_info.get("url"):
            status["upToDate"] = False
            status["reason"] = "missing-link"
        elif pdf_info.get("downloaded"):
            ref = status.get("referenceSha256")
            if ref is None:
                # First run — adopt the downloaded PDF as the reference.
                status["referenceSha256"] = stored_hash
                status["upToDate"] = True
                status.pop("reason", None)
            else:
                status["upToDate"] = stored_hash == ref
                if status["upToDate"]:
                    status.pop("reason", None)
                else:
                    status["reason"] = "pdf-changed"
        else:
            status["upToDate"] = False
            status["reason"] = "download-failed"

    doc["status"] = status
    PRICING_JSON.write_text(
        json.dumps(doc, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return status


# ------------------------------------------------------------------- CLI ---

def cmd_fetch(args):
    report = fetch_sources()
    print(json.dumps(report, ensure_ascii=False, indent=2))


def cmd_transform(args):
    write_schedules()


def cmd_pricing(args):
    # Use last known discovery if caller provides --url / --page, else just
    # refresh the stored fingerprint.
    pdf_info = None
    if args.url or args.page:
        pdf_info = {
            "url": args.url,
            "source_page": args.page or POOL_PAGES["25m"],
            "downloaded": PRICING_PDF.exists(),
        }
    status = update_pricing_status(pdf_info)
    print(json.dumps(status, ensure_ascii=False, indent=2))


def cmd_update(args):
    report = fetch_sources()
    write_schedules()
    pdf_info = {
        "url": report["pricing"].get("url"),
        "source_page": report["pricing"].get("source_page"),
        "downloaded": report["pricing"].get("downloaded", False),
    }
    status = update_pricing_status(pdf_info)
    print(json.dumps({"fetch": report, "pricing_status": status},
                     ensure_ascii=False, indent=2))


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = p.add_subparsers(dest="cmd", required=True)

    sub.add_parser("fetch", help="download XLSX + cenník PDF").set_defaults(func=cmd_fetch)
    sub.add_parser("transform", help="convert local XLSX to schedule*.json").set_defaults(func=cmd_transform)

    pp = sub.add_parser("pricing", help="refresh pricing.status")
    pp.add_argument("--url", help="current cenník PDF URL (skip discovery)")
    pp.add_argument("--page", help="pool page URL where the PDF lives")
    pp.set_defaults(func=cmd_pricing)

    sub.add_parser("update", help="fetch + transform + pricing (daily cron)").set_defaults(func=cmd_update)

    args = p.parse_args(argv)
    args.func(args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
