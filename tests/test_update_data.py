"""Golden-file tests for scripts/update_data.py.

The STARZ XLSX layout is the main fragility of this project. When it
changes, the daily pipeline stops committing `schedule*.json`. These
tests build a synthetic workbook that matches the known-good layout and
verify transform_xlsx() against a committed golden JSON, plus assert
that clear errors fire when specific landmarks (date cell,
'Počet voľných dráh' label) move.

Run with: python3 -m unittest tests.test_update_data
"""
from __future__ import annotations

import datetime as dt
import io
import json
import sys
import unittest
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import update_data  # noqa: E402

SLOTS_PER_DAY = update_data.SLOTS_PER_DAY  # 76
ROWS_PER_DAY = update_data.ROWS_PER_DAY    # 9
COUNTER_OFFSET = update_data.COUNTER_ROW_OFFSET  # 8
DATA_COL_START = update_data.DATA_COL_START  # 4


def _blank_row(cols=DATA_COL_START + SLOTS_PER_DAY):
    return [None] * cols


def build_workbook(days: list[tuple[dt.date, list[int]]]) -> bytes:
    """Build an XLSX workbook shaped like the STARZ Verejnost sheet.

    Layout (0-indexed row numbers):
      rows 0-4: header rows (dummy content, skipped by the scraper)
      row  5 + k*9:        date cell in column A
      rows 6..12 + k*9:    per-service rows (dummy)
      row 13 + k*9:        'Počet voľných dráh' label in A, lane counts from col E
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verejnost"
    # Dummy header rows 1..5 (0-indexed 0..4).
    for r in range(5):
        ws.append(["Header"] * 3)
    for day_idx, (date, lanes) in enumerate(days):
        base_row = 5 + day_idx * ROWS_PER_DAY  # 0-based row of the date
        row_date = _blank_row()
        row_date[0] = dt.datetime(date.year, date.month, date.day)
        # Pad per-service rows.
        service_rows = [_blank_row() for _ in range(COUNTER_OFFSET - 1)]
        row_counter = _blank_row()
        row_counter[0] = "Počet voľných dráh"
        for i, v in enumerate(lanes):
            row_counter[DATA_COL_START + i] = v
        for i, row in enumerate([row_date] + service_rows + [row_counter]):
            target = base_row + i + 1  # openpyxl is 1-based
            for col, val in enumerate(row, start=1):
                if val is not None:
                    ws.cell(row=target, column=col, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TransformXlsxTests(unittest.TestCase):
    def test_golden_three_days(self):
        """transform_xlsx produces a stable JSON shape for a known input."""
        days_in = [
            (dt.date(2026, 4, 20), [0] * 20 + [3] * 8 + [0] * 48),
            (dt.date(2026, 4, 21), [0] * 24 + [4] * 12 + [0] * 40),
            (dt.date(2026, 4, 22), [0] * SLOTS_PER_DAY),
        ]
        wb_bytes = build_workbook(days_in)
        xlsx_path = ROOT / "tests" / "_tmp_test.xlsx"
        xlsx_path.write_bytes(wb_bytes)
        try:
            meta = {"name": "Test Pool 25 m", "max_lanes": 4}
            out = update_data.transform_xlsx(
                xlsx_path, meta, "https://example.test/page"
            )
        finally:
            xlsx_path.unlink(missing_ok=True)
        self.assertEqual(out["pool"], "Test Pool 25 m")
        self.assertEqual(out["source"], "https://example.test/page")
        self.assertEqual(out["slotMinutes"], 15)
        self.assertEqual(out["maxLanes"], 4)
        self.assertEqual(out["dayStart"], "05:00")
        self.assertEqual(out["dayEnd"], "24:00")
        self.assertEqual(len(out["days"]), 3)
        self.assertEqual(out["days"][0]["date"], "2026-04-20")
        self.assertEqual(out["days"][0]["weekday"], "pondelok")
        self.assertEqual(len(out["days"][0]["free"]), SLOTS_PER_DAY)
        self.assertEqual(out["days"][0]["free"][:20], [0] * 20)
        self.assertEqual(out["days"][0]["free"][20:28], [3] * 8)
        self.assertEqual(out["days"][1]["weekday"], "utorok")
        self.assertEqual(out["days"][1]["free"][24:36], [4] * 12)
        self.assertEqual(out["days"][2]["free"], [0] * SLOTS_PER_DAY)

    def test_counter_row_mislabeled_raises(self):
        """If 'Počet voľných dráh' label moves, error should be explicit."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Verejnost"
        for _ in range(5):
            ws.append(["Header"])
        # Date row at position 6 (1-based) = index 5 (0-based).
        ws.cell(row=6, column=1, value=dt.datetime(2026, 4, 20))
        # Counter row at row 14 — but with wrong label.
        ws.cell(row=14, column=1, value="Voľné (zmena layoutu)")
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_path = ROOT / "tests" / "_tmp_bad.xlsx"
        xlsx_path.write_bytes(buf.getvalue())
        try:
            meta = {"name": "Test", "max_lanes": 4}
            with self.assertRaisesRegex(
                RuntimeError, "expected 'Počet voľných dráh'"
            ):
                update_data.transform_xlsx(
                    xlsx_path, meta, "https://example.test/"
                )
        finally:
            xlsx_path.unlink(missing_ok=True)

    def test_no_days_raises(self):
        """Workbook with no date cells in expected layout raises clearly."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Verejnost"
        for _ in range(5):
            ws.append(["Header"])
        ws.cell(row=6, column=1, value="not a date")
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_path = ROOT / "tests" / "_tmp_empty.xlsx"
        xlsx_path.write_bytes(buf.getvalue())
        try:
            meta = {"name": "Test", "max_lanes": 4}
            with self.assertRaisesRegex(RuntimeError, "No day rows found"):
                update_data.transform_xlsx(
                    xlsx_path, meta, "https://example.test/"
                )
        finally:
            xlsx_path.unlink(missing_ok=True)

    def test_truncated_workbook_raises(self):
        """If counter row is missing entirely (truncated), fail explicitly."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Verejnost"
        for _ in range(5):
            ws.append(["Header"])
        ws.cell(row=6, column=1, value=dt.datetime(2026, 4, 20))
        # Do NOT add the counter row — sheet ends before index+8.
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_path = ROOT / "tests" / "_tmp_trunc.xlsx"
        xlsx_path.write_bytes(buf.getvalue())
        try:
            meta = {"name": "Test", "max_lanes": 4}
            with self.assertRaisesRegex(
                RuntimeError, "(Workbook truncated|expected 'Počet)"
            ):
                update_data.transform_xlsx(
                    xlsx_path, meta, "https://example.test/"
                )
        finally:
            xlsx_path.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
